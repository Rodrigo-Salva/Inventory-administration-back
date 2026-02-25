from typing import List, Optional
from fastapi import APIRouter, Depends, Query, status
from fastapi.responses import StreamingResponse
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from ...models.base import get_db
from ...dependencies import get_current_tenant
from ...repositories.report_repo import ReportRepository
from ...schemas.report import InventoryReport
from ...core.logging_config import get_logger

logger = get_logger(__name__)
router = APIRouter()

@router.get("/dashboard", response_model=InventoryReport)
async def get_dashboard_data(
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Obtiene todos los datos necesarios para el dashboard inicial con filtros de fecha"""
    repo = ReportRepository(db)
    
    stats = await repo.get_dashboard_stats(tenant_id, start_date=start_date, end_date=end_date)
    sales_stats = await repo.get_sales_stats(tenant_id, start_date=start_date, end_date=end_date)
    stats.update({
        "sales_count": sales_stats["sales_count"],
        "total_revenue": sales_stats["total_revenue"]
    })

    trends = await repo.get_movement_trends(tenant_id, start_date=start_date, end_date=end_date)
    sales_trends = await repo.get_sales_trends(tenant_id)
    top_selling = await repo.get_top_selling_products(tenant_id)
    
    recent = await repo.get_recent_movements(tenant_id)
    low_stock = await repo.get_low_stock_products(tenant_id)
    cat_distribution = await repo.get_category_distribution(tenant_id)
    sup_distribution = await repo.get_supplier_distribution(tenant_id)
    user_activity = await repo.get_user_activity(tenant_id, start_date=start_date, end_date=end_date)
    top_moves = await repo.get_top_moving_products(tenant_id, start_date=start_date, end_date=end_date)
    
    return {
        "stats": stats,
        "trends": trends,
        "sales_trends": sales_trends,
        "top_selling_products": top_selling,
        "recent_movements": recent,
        "low_stock_products": low_stock,
        "category_distribution": cat_distribution,
        "supplier_distribution": sup_distribution,
        "user_activity": user_activity,
        "top_moving_products": top_moves
    }

@router.get("/inventory-csv")
async def export_inventory_csv(
    category_id: Optional[int] = Query(None),
    supplier_id: Optional[int] = Query(None),
    is_active: Optional[bool] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Genera un reporte CSV del inventario actual con filtros opcionales"""
    from ...repositories.product_repo import ProductRepository
    repo = ProductRepository(db)
    
    # Obtener productos filtrados
    products, _ = await repo.get_filtered(
        tenant_id=tenant_id,
        category_id=category_id,
        supplier_id=supplier_id,
        is_active=is_active
    )
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["ID", "SKU", "Nombre", "Categoría", "Stock", "Precio", "Costo", "Estado"])
    
    for p in products:
        writer.writerow([
            p.id, 
            p.sku, 
            p.name, 
            p.category.name if p.category else "", 
            p.stock, 
            p.price, 
            p.cost, 
            "Activo" if p.is_active else "Inactivo"
        ])
    
    filename = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/inventory-excel")
async def export_inventory_excel(
    category_id: Optional[int] = Query(None),
    supplier_id: Optional[int] = Query(None),
    is_active: Optional[bool] = Query(None),
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Genera un reporte Excel profesional del inventario con filtros de fecha y más"""
    from ...repositories.product_repo import ProductRepository
    from ...repositories.tenant_repo import TenantRepository
    repo = ProductRepository(db)
    t_repo = TenantRepository(db)
    
    products, _ = await repo.get_filtered(
        tenant_id=tenant_id,
        category_id=category_id,
        supplier_id=supplier_id,
        is_active=is_active,
        start_date=start_date,
        end_date=end_date
    )
    tenant = await t_repo.get_by_id(tenant_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos profesionales
    title_font = Font(name='Arial', size=16, bold=True, color="1E40AF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color="DDDDDD"),
        right=Side(style='thin', color="DDDDDD"),
        top=Side(style='thin', color="DDDDDD"),
        bottom=Side(style='thin', color="DDDDDD")
    )

    # Header Corporativo
    ws.merge_cells('A1:I1')
    tenant_name = tenant.name.upper() if tenant else "SISTEMA"
    ws['A1'] = f"REPORTE DE PRODUCTOS - {tenant_name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:I2')
    date_range_str = f"Rango: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}" if start_date and end_date else "Todos los registros"
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | {date_range_str}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ["ID", "SKU", "Nombre", "Categoría", "Stock", "Mínimo", "Precio", "Costo", "Estado"]
    ws.append([])
    ws.append(headers)
    
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for idx, p in enumerate(products, start=5):
        row = [p.id, p.sku, p.name, p.category.name if p.category else "Sin categoría", p.stock, p.min_stock, float(p.price), float(p.cost or 0), "ACTIVO" if p.is_active else "INACTIVO"]
        ws.append(row)
        for cell in ws[idx]:
            cell.border = thin_border
            if cell.column in [7, 8]: cell.number_format = '"$"#,##0.00'

    # Auto-ajustar columnas
    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 15

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return StreamingResponse(excel_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Productos_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/suppliers-excel")
async def export_suppliers_excel(
    search: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    from ...repositories.supplier_repo import SupplierRepository
    from ...repositories.tenant_repo import TenantRepository
    repo = SupplierRepository(db)
    t_repo = TenantRepository(db)
    suppliers, _ = await repo.get_filtered(tenant_id=tenant_id, search=search, is_active=is_active, start_date=start_date, end_date=end_date)
    tenant = await t_repo.get_by_id(tenant_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    title_font = Font(name='Arial', size=16, bold=True, color="1E40AF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="DDDDDD"), right=Side(style='thin', color="DDDDDD"), top=Side(style='thin', color="DDDDDD"), bottom=Side(style='thin', color="DDDDDD"))

    ws.merge_cells('A1:H1')
    tenant_name = tenant.name.upper() if tenant else "SISTEMA"
    ws['A1'] = f"REPORTE DE PROVEEDORES - {tenant_name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:H2')
    date_range_str = f"Rango: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}" if start_date and end_date else "Todos los registros"
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | {date_range_str}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ["ID", "Código", "Nombre", "Email", "Teléfono", "RUC", "Estado", "Fecha Registro"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for idx, s in enumerate(suppliers, start=5):
        row = [s.id, s.code, s.name, s.email, s.phone, s.tax_id, "ACTIVO" if s.is_active else "INACTIVO", s.created_at.strftime('%d/%m/%Y %H:%M')]
        ws.append(row)
        for cell in ws[idx]: cell.border = thin_border

    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return StreamingResponse(excel_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Proveedores_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/categories-excel")
async def export_categories_excel(
    search: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    from ...repositories.category_repo import CategoryRepository
    from ...repositories.tenant_repo import TenantRepository
    repo = CategoryRepository(db)
    t_repo = TenantRepository(db)
    categories, _ = await repo.get_filtered(tenant_id=tenant_id, search=search, is_active=is_active, start_date=start_date, end_date=end_date)
    tenant = await t_repo.get_by_id(tenant_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    
    title_font = Font(name='Arial', size=16, bold=True, color="1E40AF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="DDDDDD"), right=Side(style='thin', color="DDDDDD"), top=Side(style='thin', color="DDDDDD"), bottom=Side(style='thin', color="DDDDDD"))

    ws.merge_cells('A1:F1')
    tenant_name = tenant.name.upper() if tenant else "SISTEMA"
    ws['A1'] = f"REPORTE DE CATEGORÍAS - {tenant_name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:F2')
    date_range_str = f"Rango: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}" if start_date and end_date else "Todos los registros"
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | {date_range_str}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ["ID", "Código", "Nombre", "Descripción", "Estado", "Fecha Registro"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for idx, c in enumerate(categories, start=5):
        row = [c.id, c.code, c.name, c.description, "ACTIVO" if c.is_active else "INACTIVO", c.created_at.strftime('%d/%m/%Y %H:%M')]
        ws.append(row)
        for cell in ws[idx]: cell.border = thin_border

    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return StreamingResponse(excel_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Categorias_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/users-excel")
async def export_users_excel(
    search: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    from ...repositories.user_repo import UserRepository
    from ...repositories.tenant_repo import TenantRepository
    repo = UserRepository(db)
    t_repo = TenantRepository(db)
    users, _ = await repo.get_filtered(tenant_id=tenant_id, search=search, is_active=is_active, start_date=start_date, end_date=end_date)
    tenant = await t_repo.get_by_id(tenant_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    title_font = Font(name='Arial', size=16, bold=True, color="1E40AF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="DDDDDD"), right=Side(style='thin', color="DDDDDD"), top=Side(style='thin', color="DDDDDD"), bottom=Side(style='thin', color="DDDDDD"))

    ws.merge_cells('A1:E1')
    tenant_name = tenant.name.upper() if tenant else "SISTEMA"
    ws['A1'] = f"REPORTE DE USUARIOS - {tenant_name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:E2')
    date_range_str = f"Rango: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}" if start_date and end_date else "Todos los registros"
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | {date_range_str}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ["ID", "Email", "Tipo", "Estado", "Fecha Registro"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for idx, u in enumerate(users, start=5):
        row = [u.id, u.email, "ADMIN" if u.is_admin else "OPERADOR", "ACTIVO" if u.is_active else "INACTIVO", u.created_at.strftime('%d/%m/%Y %H:%M')]
        ws.append(row)
        for cell in ws[idx]: cell.border = thin_border

    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return StreamingResponse(excel_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Usuarios_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/movements-excel")
async def export_movements_excel(
    product_id: Optional[int] = Query(None),
    movement_type: Optional[str] = Query(None),
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    from ...repositories.inventory_movement_repo import InventoryMovementRepository
    from ...repositories.tenant_repo import TenantRepository
    repo = InventoryMovementRepository(db)
    t_repo = TenantRepository(db)
    movements, _ = await repo.get_filtered(tenant_id=tenant_id, product_id=product_id, movement_type=movement_type, start_date=start_date, end_date=end_date)
    tenant = await t_repo.get_by_id(tenant_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    title_font = Font(name='Arial', size=16, bold=True, color="1E40AF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="DDDDDD"), right=Side(style='thin', color="DDDDDD"), top=Side(style='thin', color="DDDDDD"), bottom=Side(style='thin', color="DDDDDD"))

    ws.merge_cells('A1:I1')
    tenant_name = tenant.name.upper() if tenant else "SISTEMA"
    ws['A1'] = f"KARDEX DE MOVIMIENTOS - {tenant_name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:I2')
    date_range_str = f"Rango: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}" if start_date and end_date else "Todos los registros"
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | {date_range_str}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ["ID", "Fecha", "Producto", "Tipo", "Cantidad", "Stock Anterior", "Stock Nuevo", "Referencia", "Notas"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for idx, m in enumerate(movements, start=5):
        row = [m.id, m.created_at.strftime('%d/%m/%Y %H:%M'), m.product.name if m.product else "N/A", m.movement_type.upper(), m.quantity, m.stock_before, m.stock_after, m.reference, m.notes]
        ws.append(row)
        for cell in ws[idx]: cell.border = thin_border

    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return StreamingResponse(excel_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Kardex_{datetime.now().strftime('%Y%m%d')}.xlsx"})

@router.get("/sales-excel")
async def export_sales_excel(
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    status: Optional[str] = Query(None),
    payment_method: Optional[str] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Genera un reporte Excel profesional de ventas"""
    from ...repositories.tenant_repo import TenantRepository
    repo = ReportRepository(db)
    t_repo = TenantRepository(db)
    
    sales = await repo.get_filtered_sales(
        tenant_id=tenant_id,
        start_date=start_date,
        end_date=end_date,
        status=status,
        payment_method=payment_method
    )
    tenant = await t_repo.get_by_id(tenant_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    title_font = Font(name='Arial', size=16, bold=True, color="1E40AF")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color="DDDDDD"), right=Side(style='thin', color="DDDDDD"), top=Side(style='thin', color="DDDDDD"), bottom=Side(style='thin', color="DDDDDD"))

    ws.merge_cells('A1:G1')
    tenant_name = tenant.name.upper() if tenant else "SISTEMA"
    ws['A1'] = f"REPORTE DE VENTAS - {tenant_name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    date_range_str = f"Rango: {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}" if start_date and end_date else "Historial Completo"
    ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | {date_range_str}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ["ID", "Fecha/Hora", "Método", "Vendedor", "Estado", "Items", "Total"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for idx, s in enumerate(sales, start=5):
        items_summary = ", ".join([f"{item.product.name} (x{item.quantity})" for item in s.items if item.product])
        row = [
            s.id, 
            s.created_at.strftime('%d/%m/%Y %H:%M'), 
            s.payment_method.upper(), 
            s.user.email.split('@')[0] if s.user else "N/A",
            s.status.upper(),
            items_summary,
            float(s.total_amount)
        ]
        ws.append(row)
        for cell in ws[idx]: 
            cell.border = thin_border
            if cell.column == 7: cell.number_format = '"$"#,##0.00'

    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 20
    ws.column_dimensions['F'].width = 50 # Items summary wider

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return StreamingResponse(
        excel_file, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=Ventas_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@router.get("/sales-pdf")
async def export_sales_pdf(
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    status: Optional[str] = Query(None),
    payment_method: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Genera un reporte PDF resumen de ventas"""
    from ...repositories.tenant_repo import TenantRepository
    from ...services.report_generator import ReportGenerator
    # Usamos el repo de reportes para obtener las ventas filtradas
    repo = ReportRepository(db)
    t_repo = TenantRepository(db)
    
    # Nota: He actualizado el repo de reportes para soportar 'search' en la consulta si es necesario, 
    # pero para el reporte PDF usaremos una lógica similar al excel.
    # Primero obtenemos las ventas.
    from ...repositories.sale_repo import SaleRepository
    s_repo = SaleRepository(db)
    
    # Obtenemos todas las ventas filtradas (sin paginación para el reporte completo)
    # Reutilizamos la lógica del repo de ventas pero sin offset/limit
    sales, _ = await s_repo.get_sales_paginated(
        tenant_id=tenant_id,
        page=1,
        size=500, # Un límite razonable para el PDF
        start_date=start_date,
        end_date=end_date,
        status=status,
        payment_method=payment_method,
        search=search
    )
    
    tenant = await t_repo.get_by_id(tenant_id)
    tenant_name = tenant.name if tenant else "Mi Negocio"
    
    filters = {
        "start_date": start_date,
        "end_date": end_date,
        "status": status,
        "payment_method": payment_method,
        "search": search
    }
    
    pdf_buffer = ReportGenerator.generate_sales_summary_pdf(sales, tenant_name, filters)
    
    filename = f"Reporte_Ventas_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/sales-history-stats")
async def get_sales_history_stats(
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    status: Optional[str] = Query(None),
    payment_method: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    seller_id: Optional[int] = Query(None),
    tenant_id: int = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Obtiene KPIs y tendencias para la página de historial de ventas"""
    repo = ReportRepository(db)
    return await repo.get_sales_history_stats(
        tenant_id=tenant_id,
        start_date=start_date,
        end_date=end_date,
        status=status,
        payment_method=payment_method,
        search=search,
        seller_id=seller_id
    )
